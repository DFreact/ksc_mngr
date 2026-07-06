#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генерирует каталоги (YAML) для приложения управления KSC из исходной Excel-таблицы.
Запуск:  python3 build_catalogs.py <путь_к_xlsx> <выходная_папка_catalogs>
Идемпотентно: перезапускай после обновления Excel, каталоги перегенерируются.
"""
import sys, re, os
import openpyxl
import yaml

SRC = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/Политики_и_настройки_KSC_anon.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/home/claude/catalogs"
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)

def slug(s):
    s = (s or "").strip().lower()
    s = re.sub(r"[^\w\s-]", "", s, flags=re.U)
    s = re.sub(r"[\s-]+", "_", s, flags=re.U)
    return s.strip("_") or "x"

def dump(obj, fname):
    p = os.path.join(OUT, fname)
    with open(p, "w", encoding="utf-8") as f:
        yaml.safe_dump(obj, f, allow_unicode=True, sort_keys=False, width=120)
    print(f"  → {fname}: записано")

YES = {"да", "включено", "включен", "вкл"}
NO = {"нет", "выключено", "выключен", "выкл", "отключено"}
NA = {"не применимо", "не доступно", "недоступно", None, ""}

def norm(v):
    return str(v).strip() if v is not None else ""

def infer_type(values):
    """values — список сырых значений параметра по всем политикам (без 'не применимо')."""
    vals = [norm(v) for v in values if norm(v).lower() not in NA]
    if not vals:
        return "bool", None
    low = {v.lower() for v in vals}
    if low <= (YES | NO):
        return "bool", None
    # число (в т.ч. с единицей вроде 3600)
    if all(re.fullmatch(r"-?\d+(?:[.,]\d+)?", v) for v in vals):
        return "number", None
    # небольшой набор уникальных строк → enum
    uniq = sorted(set(vals))
    if len(uniq) <= 8:
        return "enum", uniq
    return "string", None


# ─────────────────────────────────────────────────────────────────────────────
# 1. МАТРИЦА ПОЛИТИК KESL  →  параметры + значения политик
# ─────────────────────────────────────────────────────────────────────────────
def parse_kesl_matrix():
    ws = wb["Политики KESL"]
    # колонки политик: (код, столбец_принудительно, столбец_значения)
    policies = [
        ("MAIN", 6, 7), ("PD", 8, 9), ("VD", 10, 11), ("HSPD", 12, 13),
        ("SPD", 14, 15), ("SVD", 16, 17), ("SVDD", 18, 19), ("SVDWS", 20, 21),
        ("SVDFS", 22, 23), ("SVDDC", 24, 25), ("HSVD", 26, 27),
    ]
    params = []
    policy_values = {code: [] for code, _, _ in policies}
    cat = subcat = grp = ""
    seen = {}
    for r in range(3, ws.max_row + 1):
        num = ws.cell(r, 1).value
        c = norm(ws.cell(r, 2).value)
        s = norm(ws.cell(r, 3).value)
        g = norm(ws.cell(r, 4).value)
        name = norm(ws.cell(r, 5).value)
        if c: cat = c
        if s: subcat = s
        if g: grp = g
        if not name:
            continue
        # пропускаем «шапочные» служебные строки политики
        if name in ("Название", "Целевая группа администрирования", "Приложение",
                    "Состояние политики"):
            continue
        raw_vals = [ws.cell(r, col).value for _, _, col in policies]
        vtype, opts = infer_type(raw_vals)
        pid = f"kesl.{slug(cat)}.{slug(subcat)}.{slug(grp)}.{slug(name)}"
        base = pid
        i = 2
        while pid in seen:
            pid = f"{base}_{i}"; i += 1
        seen[pid] = True
        entry = {
            "id": pid,
            "application": "kesl",
            "category": cat,
            "subcategory": subcat or None,
            "group": grp or None,
            "name": name,
            "value_type": vtype,
            "added_in_version": "12.4",   # правится вручную при верификации других версий
        }
        if opts:
            entry["enum_options"] = opts
        # плюсы/минусы и покрытие атак — пустые заготовки под наполнение
        entry["tradeoffs_pros"] = []
        entry["tradeoffs_cons"] = []
        entry["tradeoffs_reviewed"] = False
        params.append(entry)
        for code, fcol, vcol in policies:
            fv = norm(ws.cell(r, fcol).value).lower()
            forced = True if fv in YES else (False if fv in NO else None)
            policy_values[code].append({
                "parameter_id": pid,
                "value": norm(ws.cell(r, vcol).value),
                "forced": forced,
            })
    dump(params, "kesl_parameters.yaml")
    dump({"policies": policy_values}, "kesl_policy_values.yaml")
    return len(params)


# ─────────────────────────────────────────────────────────────────────────────
# 2. КОНТРОЛЬ УСТРОЙСТВ
# ─────────────────────────────────────────────────────────────────────────────
def parse_device_control():
    ws = wb["KESL Контроль Устройств"]
    device_types = {}   # имя → набор режимов доступа
    section = ""
    access_options = set()
    for r in range(1, ws.max_row + 1):
        a = norm(ws.cell(r, 1).value)
        b = norm(ws.cell(r, 2).value)
        if not a:
            continue
        # секции без второй колонки — заголовки групп устройств
        if a and not b and a not in ("Имя",):
            if a in ("Устройства и сети Wi-Fi", "Внешние устройства", "Сети Wi-Fi",
                     "Прочие устройства", "Пользовательские правила", "Шины подключения"):
                section = a
                continue
        if a in ("Запоминающие устройства", "Имя"):  # строки-заголовки таблиц
            continue
        if section in ("Устройства и сети Wi-Fi", "Внешние устройства", "Сети Wi-Fi",
                       "Прочие устройства") and b:
            device_types.setdefault(a, {"section": section, "access_options": []})
            if b not in device_types[a]["access_options"]:
                device_types[a]["access_options"].append(b)
            access_options.add(b)
    catalog = {
        "device_types": [
            {"id": slug(k), "name": k, "section": v["section"],
             "access_options": v["access_options"]}
            for k, v in device_types.items()
        ],
        "all_access_options": sorted(access_options),
        # схема таблицы пользовательских правил доступа (как в KSC)
        "custom_rules_schema": [
            {"key": "name", "label": "Имя", "type": "string"},
            {"key": "device_type", "label": "Тип устройства", "type": "string"},
            {"key": "priority", "label": "Приоритет", "type": "number"},
            {"key": "user_or_group", "label": "Пользователь или группа", "type": "string"},
            {"key": "comment", "label": "Комментарий", "type": "string"},
            {"key": "schedule_status", "label": "Статус расписания доступа", "type": "bool"},
            {"key": "schedule", "label": "Расписание доступа", "type": "string"},
            {"key": "read", "label": "Чтение", "type": "bool"},
            {"key": "write", "label": "Запись", "type": "bool"},
        ],
        "bus_types": ["USB", "FireWire"],
        "bus_access_options": ["Разрешать", "Запрещать"],
    }
    dump(catalog, "device_control.yaml")
    return len(catalog["device_types"])


# ─────────────────────────────────────────────────────────────────────────────
# 3. СХЕМЫ ВЛОЖЕННЫХ СПИСКОВ-ТАБЛИЦ  (то, чего не хватает в приложении)
#    Схемы — стабильная структура KSC; сид-строки берём с листа KESL MAIN.
# ─────────────────────────────────────────────────────────────────────────────
LIST_SCHEMAS = {
    "protection_scope": {
        "name": "Область защиты", "component": "Защита от файловых угроз",
        "columns": [
            {"key": "scope", "label": "Область защиты", "type": "string"},
            {"key": "status", "label": "Статус", "type": "enum", "options": ["Включено", "Выключено"]},
        ]},
    "malware_exclusions": {
        "name": "Исключения из поиска вредоносного ПО", "component": "Защита от файловых угроз",
        "columns": [
            {"key": "name", "label": "Имя исключения", "type": "string"},
            {"key": "file_or_folder", "label": "Файл или папка", "type": "string"},
            {"key": "object_name", "label": "Имя объекта", "type": "string"},
            {"key": "component", "label": "Компонент защиты", "type": "string"},
        ]},
    "trusted_apps": {
        "name": "Доверенные приложения", "component": "Анализ поведения",
        "columns": [
            {"key": "application", "label": "Приложение", "type": "string"},
            {"key": "path", "label": "Путь", "type": "string"},
            {"key": "comment", "label": "Комментарий", "type": "string"},
        ]},
    "attachment_types": {
        "name": "Типы вложений", "component": "Защита от почтовых угроз",
        "columns": [
            {"key": "status", "label": "Статус", "type": "enum", "options": ["Включен", "Выключен"]},
            {"key": "mask", "label": "Маска файла вложения", "type": "string"},
        ]},
    "trusted_urls": {
        "name": "Доверенные веб-адреса", "component": "Защита от веб-угроз",
        "columns": [
            {"key": "url", "label": "Веб-адрес", "type": "string"},
        ]},
    "trusted_ips": {
        "name": "Доверенные IP-адреса", "component": "Защита от сетевых угроз",
        "columns": [
            {"key": "ip", "label": "IP-адрес", "type": "string"},
        ]},
    "network_packet_rules": {
        "name": "Сетевые пакетные правила", "component": "Сетевой экран",
        "columns": [
            {"key": "priority", "label": "Приоритет", "type": "string"},
            {"key": "rule", "label": "Сетевое правило", "type": "string"},
            {"key": "action", "label": "Действие", "type": "enum", "options": ["Разрешить", "Запретить"]},
            {"key": "address", "label": "Адрес", "type": "string"},
            {"key": "source", "label": "Источник", "type": "string"},
        ]},
    "available_networks": {
        "name": "Доступные сети", "component": "Сетевой экран",
        "columns": [
            {"key": "name", "label": "Имя", "type": "string"},
            {"key": "net_type", "label": "Тип сети", "type": "string"},
            {"key": "ip", "label": "IP-адрес", "type": "string"},
        ]},
    "trusted_devices": {
        "name": "Доверенные устройства", "component": "Контроль устройств",
        "columns": [
            {"key": "device_name", "label": "Имя устройства", "type": "string"},
            {"key": "device_type", "label": "Тип устройства", "type": "string"},
            {"key": "device_model", "label": "Модель устройства", "type": "string"},
            {"key": "device_id", "label": "Идентификатор устройства", "type": "string"},
            {"key": "user_or_group", "label": "Пользователь или группа", "type": "string"},
            {"key": "comment", "label": "Комментарий", "type": "string"},
        ]},
    "message_templates": {
        "name": "Шаблоны сообщений", "component": "Веб-Контроль",
        "columns": [
            {"key": "attention", "label": "Внимание", "type": "text"},
            {"key": "block_message", "label": "Сообщение о блокировке", "type": "text"},
            {"key": "admin_message", "label": "Сообщение администратору", "type": "text"},
        ]},
}

def parse_list_schemas_and_seed():
    schemas = [{"id": k, **v} for k, v in LIST_SCHEMAS.items()]
    dump({"list_types": schemas}, "kesl_list_schemas.yaml")

    # Сид-строки для листа MAIN (пример наполнения; остальные политики — по аналогии).
    ws = wb["KESL MAIN"]
    seed = {}
    # блоки заданы стартовой колонкой в шапке (строка 3 = подписи колонок)
    # ищем колонки по известным подписям в строке 3
    header = {c: norm(ws.cell(3, c).value) for c in range(1, ws.max_column + 1)}
    def col_of(label):
        for c, v in header.items():
            if v == label:
                return c
        return None
    # исключения
    def collect(start_col, keys):
        rows = []
        for r in range(4, ws.max_row + 1):
            rec = {}
            empty = True
            for off, key in enumerate(keys):
                val = norm(ws.cell(r, start_col + off).value)
                if val:
                    empty = False
                rec[key] = val
            if not empty:
                rows.append(rec)
        return rows
    c = col_of("Имя исключения")
    if c:
        seed["malware_exclusions"] = collect(c, ["name", "file_or_folder", "object_name", "component"])
    c = col_of("Маска файла вложения")
    if c:
        seed["attachment_types"] = [{"mask": r["mask"]} for r in collect(c-1, ["status", "mask"]) if r.get("mask")]
    c = col_of("IP-адрес")
    if c:
        seed["trusted_ips"] = [r for r in collect(c, ["ip"]) if r.get("ip")]
    dump({"policy": "MAIN", "seed_rows": seed}, "kesl_list_seed_main.yaml")
    return len(schemas)


# ─────────────────────────────────────────────────────────────────────────────
# 4. АГЕНТ АДМИНИСТРИРОВАНИЯ  →  параметры
# ─────────────────────────────────────────────────────────────────────────────
def parse_network_agent():
    ws = wb["Политики администрирования(бс)"]
    params = []
    cat = subcat = grp = ""
    seen = {}
    # колонки значений политик начинаются с 6
    pol_cols = list(range(6, ws.max_column + 1))
    pol_names = [norm(ws.cell(1, c).value) for c in pol_cols]
    for r in range(2, ws.max_row + 1):
        c = norm(ws.cell(r, 1).value)
        s = norm(ws.cell(r, 2).value)
        g = norm(ws.cell(r, 3).value)
        name = norm(ws.cell(r, 4).value)
        if c: cat = c
        if s: subcat = s
        if g: grp = g
        if not name or name in ("Название", "Целевая группа администрирования",
                                "Приложение", "Состояние политики"):
            continue
        raw_vals = [ws.cell(r, col).value for col in pol_cols]
        vtype, opts = infer_type(raw_vals)
        pid = f"agent.{slug(cat)}.{slug(subcat)}.{slug(grp)}.{slug(name)}"
        base = pid; i = 2
        while pid in seen:
            pid = f"{base}_{i}"; i += 1
        seen[pid] = True
        e = {"id": pid, "application": "network_agent", "category": cat,
             "subcategory": subcat or None, "group": grp or None, "name": name,
             "value_type": vtype, "added_in_version": "12.4",
             "tradeoffs_pros": [], "tradeoffs_cons": [], "tradeoffs_reviewed": False}
        if opts:
            e["enum_options"] = opts
        params.append(e)
    dump(params, "network_agent_parameters.yaml")
    return len(params)


# ─────────────────────────────────────────────────────────────────────────────
# 5. СОБЫТИЯ (лист «Политики администрирования(с)»)
# ─────────────────────────────────────────────────────────────────────────────
def parse_events():
    ws = wb["Политики администрирования(с)"]
    events = []
    channels_cols = {
        "email": 4, "sms": 5, "executable": 6, "syslog": 7,
        "os_log_device": 8, "os_log_server": 9,
    }
    for r in range(2, ws.max_row + 1):
        severity = norm(ws.cell(r, 1).value)
        etype = norm(ws.cell(r, 2).value)
        if not etype:
            continue
        days = ws.cell(r, 3).value
        ev = {
            "id": slug(etype)[:60] + f"_{r}",
            "application": "network_agent",
            "name": etype,
            "severity": severity,
            "default_storage_days": int(days) if isinstance(days, (int, float)) else None,
            "channels": {k: bool(norm(ws.cell(r, c).value)) for k, c in channels_cols.items()},
        }
        events.append(ev)
    dump(events, "network_agent_events.yaml")
    return len(events)


# ─────────────────────────────────────────────────────────────────────────────
# 6. ИЕРАРХИЯ ГРУПП
# ─────────────────────────────────────────────────────────────────────────────
def parse_groups():
    """Лист содержит два блока:
      1) дерево с отступами (колонка = уровень) — только имена;
      2) таблица описаний (шапка «1 Уровень | 2 Уровень | 3 Уровень | Описание |
         Критерий добавления | политика Агента | политика KESL») — полные пути
         с метаданными.
    Основной источник — таблица описаний; она богаче и однозначна.
    """
    ws = wb["Иерархия групп"]

    # найти строку-шапку таблицы описаний
    header_row = None
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(r, 1).value) == "1 Уровень":
            header_row = r
            break

    nodes = []
    by_path = {}  # ("Управляемые устройства", "Физические устройства") → node id

    if header_row:
        for r in range(header_row + 1, ws.max_row + 1):
            levels = [norm(ws.cell(r, c).value) for c in (1, 2, 3)]
            path = tuple(v for v in levels if v)
            if not path:
                continue
            if path in by_path:   # дубль пути — пропускаем
                continue
            nid = slug(path[-1]) + f"_{r}"
            parent = by_path.get(path[:-1])
            by_path[path] = nid
            nodes.append({
                "id": nid,
                "name": path[-1],
                "parent_id": parent,
                "depth": len(path),
                "description": norm(ws.cell(r, 4).value) or None,
                "add_criterion": norm(ws.cell(r, 5).value) or None,
                "agent_policy": norm(ws.cell(r, 6).value) or None,
                "kesl_policy": norm(ws.cell(r, 7).value) or None,
            })
    else:
        # запасной путь: дерево с отступами без метаданных
        stack = {}
        for r in range(2, ws.max_row + 1):
            depth = None; name = None
            for c in range(1, 8):
                v = norm(ws.cell(r, c).value)
                if v:
                    depth = c; name = v; break
            if not name or name.startswith("Иерархия групп") or name.startswith("Описание"):
                continue
            nid = slug(name) + f"_{r}"
            parent = None
            for d in range(depth - 1, 0, -1):
                if d in stack:
                    parent = stack[d]; break
            stack[depth] = nid
            for d in list(stack):
                if d > depth:
                    del stack[d]
            nodes.append({"id": nid, "name": name, "parent_id": parent, "depth": depth,
                          "description": None, "add_criterion": None,
                          "agent_policy": None, "kesl_policy": None})

    dump({"groups": nodes}, "group_tree.yaml")
    return len(nodes)


print("Генерация каталогов из:", SRC)
n1 = parse_kesl_matrix()
n2 = parse_device_control()
n3 = parse_list_schemas_and_seed()
n4 = parse_network_agent()
n5 = parse_events()
n6 = parse_groups()
print(f"\nИтого: KESL-параметров {n1}, типов устройств {n2}, схем списков {n3}, "
      f"параметров Агента {n4}, событий {n5}, групп {n6}")
